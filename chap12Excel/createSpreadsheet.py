import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Spam Bacon Eggs Sheet'
sheet['A1'] = "Hello World"
sheet['B1'] = "Hello World"
sheet['C1'] = "Hello World"
sheet['D1'] = "Hello World"
wb.create_sheet('Sheet2')
wb.create_sheet(index=0, title='First Sheet')
wb.create_sheet(index=2, title='Hello World')
wb.create_sheet(index=3, title='Middle Sheet')
# wb.remove_sheet(sheet['Middle Sheet'])
# wb.save('example_copy.xlsx')
# print(wb.sheetnames)
