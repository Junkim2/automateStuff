import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb=openpyxl.load_workbook('example.xlsx')
sheet=wb.active
bRow = sheet[2]
for cellObj in bRow:
    print(cellObj.value)
bCol = sheet[get_column_letter(2)]
for cellObj in bCol:
    print(cellObj.value)
